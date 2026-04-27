#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
简历风险评估脚本 v3.0 - 支持PDF/Word/Excel简历文件
针对PDF提取文本质量差的情况优化,支持Windows乱码处理
"""

import sys
import os
import json
import re
from datetime import datetime
from pathlib import Path

# 尝试导入必要的库
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("警告: 未安装openpyxl,将只输出JSON结果")

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

try:
    import docx
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    pass  # 已在上面检查


# ── 风险维度定义 ──────────────────────────────────────────────────────────────

RISK_LEVELS = ['低', '中', '高']
RISK_COLORS = {'低': 'C6EFCE', '中': 'FFEB9C', '高': 'FFC7CE'}
RISK_FONTS = {'低': '006100', '中': '9C6500', '高': '9C0006'}
RISK_DIMENSION_LABELS = {
    'education': '学历',
    'work_experience': '工作经历',
    'timeline': '时间线',
    'salary': '薪资',
    'skill_exaggeration': '技能夸大',
    'vague_language': '表述模糊',
    'professional_match': '专业匹配',
}

SECTION_ALIASES = {
    "教育背景": "education",
    "教育经历": "education",
    "工作经历": "work_experience",
    "任职经历": "work_experience",
    "工作经验": "work_experience",
    "项目经历": "project_experience",
    "项目经验": "project_experience",
    "专业技能": "skills",
    "个人技能": "skills",
    "技术专长": "skills",
    "技能栈": "skills",
    "个人优势": "self_summary",
    "个人评价": "self_summary",
    "自我评价": "self_summary",
}


def risk_summary(risks):
    return ';'.join(
        f"[{RISK_DIMENSION_LABELS.get(key, key)}]{value['level']}"
        for key, value in risks.items()
        if value['flags']
    ) or '无明显风险点'


def risk_bucket():
    return {"level": "低", "flags": [], "evidence": []}


def add_risk(risks, dimension, level, flag, evidence=""):
    current = risks[dimension]["level"]
    if RISK_LEVELS.index(level) > RISK_LEVELS.index(current):
        risks[dimension]["level"] = level
    risks[dimension]["flags"].append(flag)
    risks[dimension].setdefault("evidence", []).append((evidence or "")[:500])


def matching_evidence(text, patterns, limit=3):
    if not text:
        return ""
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines() if line.strip()]
    evidence = []
    for line in lines:
        if any(re.search(pattern, line, re.IGNORECASE) for pattern in patterns):
            evidence.append(line[:180])
            if len(evidence) >= limit:
                break
    return " | ".join(evidence)


def assess_parse_quality(parsed):
    warnings = []
    raw_text = parsed.get("raw_text", "")
    work_count = len(parsed.get("work_experience", []))
    project_count = len(parsed.get("project_experience", []))
    education_count = len(parsed.get("education", []))

    if len(raw_text) < 300:
        warnings.append("文本抽取内容较短")
    if work_count == 0:
        warnings.append("未识别到工作经历")
    if education_count == 0:
        warnings.append("未识别到教育经历")
    if "项目经历" in raw_text and project_count == 0:
        warnings.append("项目经历标题存在但未成功分区")

    if len(warnings) >= 2:
        quality = "较差"
    elif warnings:
        quality = "需复核"
    else:
        quality = "正常"

    return {
        "quality": quality,
        "warnings": warnings,
        "text_length": len(raw_text),
        "work_experience_count": work_count,
        "project_experience_count": project_count,
        "education_count": education_count,
    }


def compact_heading_text(value):
    return re.sub(r"[\s:：;；、,，.。/\\|_\-]+", "", (value or "").strip())


def normalized_heading(line):
    compact = compact_heading_text(line)
    if not compact:
        return ""
    for label in SECTION_ALIASES:
        if compact == compact_heading_text(label):
            return label
    return ""


def score_extracted_text(text):
    """Score extracted PDF text and prefer the cleaner structured variant."""
    if not text or not text.strip():
        return -1

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        return -1

    short_lines = sum(1 for line in lines if len(line) <= 2)
    meaningful_lines = sum(1 for line in lines if re.search(r'[一-龥A-Za-z]{3,}', line))
    avg_len = sum(len(line) for line in lines) / len(lines)
    first_line_bonus = 50 if re.search(r'[一-龥]{2,4}', lines[0]) else 0

    return meaningful_lines * 50 + avg_len * 5 + first_line_bonus - short_lines * 10


def clean_extracted_text(text):
    """Remove repeated watermark-like noise and normalize whitespace."""
    if not text:
        return ""

    cleaned = text.replace("\r\n", "\n")
    cleaned = re.sub(r"[A-Za-z0-9_-]{24,}~~", " ", cleaned)
    cleaned = re.sub(r"(?:\n\s*){3,}", "\n\n", cleaned)
    cleaned = re.sub(r"[ \t]{2,}", " ", cleaned)
    return cleaned.strip()


def looks_like_heading(line):
    normalized = line.strip()
    if not normalized or len(normalized) > 20:
        return False
    return bool(normalized_heading(normalized))


def extract_section_lines(lines, start_keywords, end_keywords):
    collecting = False
    results = []
    start_set = {compact_heading_text(keyword) for keyword in start_keywords}
    end_set = {compact_heading_text(keyword) for keyword in end_keywords}

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        heading = normalized_heading(stripped)
        compact = compact_heading_text(heading or stripped)
        if compact in start_set:
            collecting = True
            continue
        if collecting and (compact in end_set or (heading and compact not in start_set)):
            break
        if collecting:
            results.append(stripped)
    return results


def parse_date_token(token):
    token = (token or "").strip()
    if not token:
        return None
    if token in {"\u81f3\u4eca", "\u73b0\u5728", "\u4eca"}:
        now = datetime.now()
        return now.year * 12 + now.month

    match = re.match(r"(\d{4})(?:[./-](\d{1,2}))?$", token)
    if not match:
        return None
    year = int(match.group(1))
    month = int(match.group(2) or 1)
    return year * 12 + month


def extract_date_ranges(lines):
    pattern = re.compile(
        r"(\d{4}(?:[./-]\d{1,2})?)\s*(?:--|—|–|-|~|\u81f3)\s*(\d{4}(?:[./-]\d{1,2})?|\u81f3\u4eca|\u73b0\u5728)"
    )
    ranges = []
    for line in lines:
        match = pattern.search(line)
        if not match:
            continue
        start = parse_date_token(match.group(1))
        end = parse_date_token(match.group(2))
        if start is None or end is None:
            continue
        if start > end:
            start, end = end, start
        ranges.append((start, end, line))
    return ranges


def extract_text_from_pdf(filepath):
    """Extract text from PDF and choose the cleaner result across engines."""
    candidates = []

    if HAS_PDFPLUMBER:
        try:
            text = ""
            with pdfplumber.open(filepath) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            if text.strip():
                candidates.append(text)
        except Exception as e:
            print(f"pdfplumber提取失败: {e}")

    try:
        from PyPDF2 import PdfReader
        text = ""
        reader = PdfReader(filepath)
        for page in reader.pages:
            page_text = page.extract_text() or ""
            if page_text:
                text += page_text + "\n"
        if text.strip():
            candidates.append(text)
    except Exception as e:
        print(f"PyPDF2提取失败: {e}")

    if not candidates:
        return ""

    return clean_extracted_text(max(candidates, key=score_extracted_text))


def extract_text_from_docx(filepath):
    """从Word文档提取文本"""
    if HAS_DOCX:
        try:
            doc = docx.Document(filepath)
            return "\n".join([para.text for para in doc.paragraphs])
        except Exception as e:
            print(f"docx读取失败: {e}")
    return ""


def extract_text_from_excel(filepath):
    """从Excel提取文本"""
    if HAS_OPENPYXL:
        try:
            wb = openpyxl.load_workbook(filepath)
            text = ""
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            text += str(cell.value) + " "
                    text += "\n"
            return text
        except Exception as e:
            print(f"excel读取失败: {e}")
    return ""


def extract_resume_text(filepath):
    """智能提取简历文本,自动识别文件类型"""
    ext = Path(filepath).suffix.lower()

    if ext == '.pdf':
        return extract_text_from_pdf(filepath)
    elif ext in ['.docx', '.doc']:
        return clean_extracted_text(extract_text_from_docx(filepath))
    elif ext in ['.xlsx', '.xls']:
        return clean_extracted_text(extract_text_from_excel(filepath))
    else:
        # 尝试作为文本文件读取
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return clean_extracted_text(f.read())
        except:
            try:
                with open(filepath, 'r', encoding='gbk') as f:
                    return clean_extracted_text(f.read())
            except Exception as e:
                return f""


def looks_like_work_history_line(line, date_range_pattern):
    """Return True for date lines that look like employment, not education/projects."""
    if not date_range_pattern.search(line):
        return False

    education_keywords = ("大学", "学院", "学校", "本科", "硕士", "博士", "大专", "中专", "研究生", "教育背景", "教育经历")
    if any(keyword in line for keyword in education_keywords):
        return False

    org_keywords = ("公司", "集团", "银行", "科技", "软件", "信息", "咨询", "有限", "股份", "中心", "研究院")
    job_keywords = ("工程师", "开发", "ETL", "数据", "分析师", "架构师", "运维", "产品", "经理", "顾问", "实施")
    if any(keyword in line for keyword in org_keywords) and any(keyword in line for keyword in job_keywords):
        return True

    return False


def infer_work_experience_lines(lines, date_range_pattern):
    """Fallback for PDFs where the work-history heading is missing or glued to a row."""
    inferred = []
    current_section = ""
    for line in lines:
        heading = normalized_heading(line)
        if heading:
            current_section = SECTION_ALIASES.get(heading, "")
            continue
        if current_section == "project_experience":
            continue
        if "项目" in line and current_section != "work_experience":
            continue
        if not looks_like_work_history_line(line, date_range_pattern):
            continue
        inferred.append(line[:140])
    return list(dict.fromkeys(inferred))


def looks_like_project_timeline_line(line):
    project_terms = (
        "项目", "系统", "平台", "中台", "报送", "集市", "数仓", "风控",
        "预测性", "生产数据", "经营分析", "数据整合", "数据支撑",
    )
    company_terms = ("有限公司", "股份", "集团", "科技", "软件", "信息", "咨询")
    role_terms = ("工程师", "开发", "分析师", "顾问", "经理")
    if "项目" in line:
        return True
    if any(term in line for term in project_terms) and not (
        any(term in line for term in company_terms) and any(term in line for term in role_terms) and len(line) < 45
    ):
        return True
    return False


def parse_resume(text, input_name=None, filename=None):
    """Parse resume text into structured fields used by the risk rules."""
    normalized_text = clean_extracted_text(text)
    lines = [re.sub(r"\s+", " ", line).strip() for line in normalized_text.split("\n") if line.strip()]

    name_from_file = None
    if filename:
        name_match = re.search(r"[-_]([^-_]{2,4})(?:\.[^.]+)$", filename, re.IGNORECASE)
        if name_match:
            name_from_file = name_match.group(1)

    result = {
        "name": input_name or name_from_file or "",
        "education": [],
        "work_experience": [],
        "skills": [],
        "project_experience": [],
        "raw_text": normalized_text,
        "clean_text": " ".join(lines),
        "filename": filename or "",
        "text_before_project": normalized_text,
    }

    if not result["name"]:
        name_patterns = [
            r"姓名[::\s]*([\u4e00-\u9fa5]{2,4})",
            r"^([\u4e00-\u9fa5]{2,3})\s+(?:个人简历|简历|求职)",
            r"([\u4e00-\u9fa5]{2,4})\s{1,3}\d{3}[-\d]{7,}",
            r"([\u4e00-\u9fa5]{2,4})\s{0,3}[_~\d](?:\s|$)",
        ]
        for line in lines:
            for pattern in name_patterns:
                match = re.search(pattern, line)
                if match:
                    result["name"] = match.group(1)
                    break
            if result["name"]:
                break
        if not result["name"]:
            match = re.search(r"([\u4e00-\u9fa5]{2,4})", result["clean_text"][:200])
            result["name"] = match.group(1) if match else "未知"

    age_match = re.search(r"(\d{2})\s*(?:岁|年龄)", normalized_text)
    if age_match:
        result["age"] = age_match.group(1)

    exp_match = re.search(r"(\d{1,2})\+?\s*年(?:工作|经验)", normalized_text)
    if exp_match:
        result["work_years"] = exp_match.group(1)

    if any(normalized_heading(line) in {"项目经历", "项目经验"} for line in lines):
        prefix = []
        for line in lines:
            if normalized_heading(line) in {"项目经历", "项目经验"}:
                break
            prefix.append(line)
        result["text_before_project"] = "\n".join(prefix)

    education_section = extract_section_lines(
        lines,
        ["教育背景", "教育经历"],
        ["工作经历", "工作经验", "任职经历", "项目经历", "项目经验", "专业技能", "个人技能", "技术专长", "技能栈", "个人优势", "个人评价", "自我评价"],
    )
    education_keywords = ("大学", "学院", "学校", "本科", "硕士", "博士", "大专", "中专", "研究生")
    date_range_pattern = re.compile(
        r"\d{4}(?:[./-]\d{1,2})?\s*(?:--|—|–|-|~|至)\s*(?:\d{4}(?:[./-]\d{1,2})?|至今|现在)"
    )
    education_lines = []
    for line in education_section or lines:
        if any(keyword in line for keyword in education_keywords) and (
            date_range_pattern.search(line) or "统招" in line or "学历" in line
        ):
            education_lines.append(line[:120])
    result["education"] = list(dict.fromkeys(education_lines))

    work_section = extract_section_lines(
        lines,
        ["工作经历", "工作经验", "任职经历"],
        ["项目经历", "项目经验", "专业技能", "个人技能", "技术专长", "技能栈", "教育背景", "教育经历", "个人优势", "个人评价", "自我评价"],
    )
    work_lines = []
    for line in work_section:
        if date_range_pattern.search(line):
            work_lines.append(line[:140])
            continue
        if work_lines and any(keyword in line for keyword in ("公司", "岗位", "职责", "负责", "任", "工程师", "开发")):
            work_lines.append(line[:140])
    if not work_lines:
        work_lines = infer_work_experience_lines(lines, date_range_pattern)
    result["work_experience"] = list(dict.fromkeys(work_lines))

    result["project_experience"] = extract_section_lines(
        lines,
        ["项目经历", "项目经验"],
        ["专业技能", "个人技能", "技术专长", "技能栈", "教育背景", "教育经历", "工作经历", "工作经验", "任职经历", "个人优势", "个人评价", "自我评价"],
    )

    skill_section = extract_section_lines(
        lines,
        ["专业技能", "个人技能", "技术专长", "擅长", "技能栈"],
        ["项目经历", "项目经验", "教育背景", "教育经历", "工作经历", "工作经验", "任职经历", "个人优势", "个人评价", "自我评价"],
    )
    skill_lines = []
    for line in skill_section:
        if len(line) <= 120:
            skill_lines.append(line[:120])
    result["skills"] = list(dict.fromkeys(skill_lines))
    result["parse_quality"] = assess_parse_quality(result)

    return result


def safe_debug_filename(value, fallback="resume"):
    stem = Path(value or fallback).stem
    stem = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", stem).strip(" ._")
    return stem or fallback


def write_extraction_debug_files(parsed, output_dir, index):
    """Write raw extracted text and parsed sections for manual extraction review."""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    stem = safe_debug_filename(parsed.get("filename") or parsed.get("name"), f"resume-{index}")
    prefix = f"{index:03d}-{stem}"
    text_path = output_path / f"{prefix}.txt"
    sections_path = output_path / f"{prefix}.sections.json"

    text_path.write_text(parsed.get("raw_text", ""), encoding="utf-8")
    sections = {
        "name": parsed.get("name", ""),
        "filename": parsed.get("filename", ""),
        "text_length": len(parsed.get("raw_text", "")),
        "parse_quality": parsed.get("parse_quality", {}),
        "sections": {
            "education": parsed.get("education", []),
            "work_experience": parsed.get("work_experience", []),
            "project_experience": parsed.get("project_experience", []),
            "skills": parsed.get("skills", []),
        },
    }
    sections_path.write_text(json.dumps(sections, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"text": str(text_path), "sections": str(sections_path)}


def analyze_risk(resume_data):
    """分析简历风险,返回风险评估结果"""
    risks = {
        'education': {'level': '低', 'flags': []},
        'work_experience': {'level': '低', 'flags': []},
        'timeline': {'level': '低', 'flags': []},
        'salary': {'level': '低', 'flags': []},
        'skill_exaggeration': {'level': '低', 'flags': []},
        'vague_language': {'level': '低', 'flags': []},
        'professional_match': {'level': '低', 'flags': []}  # 新增:专业匹配度
    }

    text = resume_data.get('clean_text', '')
    work_exp = resume_data.get('work_experience', [])
    edu = resume_data.get('education', [])
    name = resume_data.get('name', '')
    raw_text = resume_data.get('raw_text', '')
    skills = resume_data.get('skills', [])

    # ── 专业匹配度分析(数据开发岗位)──
    non_cs_majors = ['会计', '财务', '市场', '营销', '材料', '高分子', '工艺', '工商']
    for major in non_cs_majors:
        if major in text[:500] or major in ' '.join(edu)[:300]:
            risks['professional_match']['level'] = '中'
            risks['professional_match']['flags'].append(f'学历专业为"{major}",与数据开发岗位关联度较低')
            break

    cs_majors = ['计算机', '软件', '信息', '大数据', '统计', '数学']
    has_cs_bg = any(m in text[:500] or m in ' '.join(edu)[:300] for m in cs_majors)

    if not has_cs_bg and risks['professional_match']['level'] == '低':
        # 未识别到明确专业背景
        pass
    elif has_cs_bg:
        risks['professional_match']['level'] = '低'
        risks['professional_match']['flags'] = []

    # ── 时间线分析(改进版,减少误报)──
    year_pattern = r'(?:20)1[5-9]|20[2-3]\d'  # 只匹配合理的年份
    years = re.findall(year_pattern, text)
    years = sorted(set([int(y) for y in years]))

    if len(years) >= 2:
        year_counts = {}
        for y in years:
            year_counts[y] = text.count(str(y))

        # 只在年数 >= 4 且有明显重叠时报警
        if len(years) >= 4:
            max_count = max(year_counts.values())
            min_count = min(year_counts.values())
            if max_count > 3 and max_count / min_count > 2:
                risks['timeline']['level'] = '中'
                risks['timeline']['flags'].append(f'工作时间段可能存在重叠,年份出现频率异常')

    # ── 职位/公司模糊分析 ─
    vague_patterns = [
        (r'某[知名大型热门]\w*', '包含"某知名/大型"等模糊公司描述'),
        (r'\w{2,4}互联网\w{0,4}', '公司描述过于泛化'),
        (r'重要项目|核心产品|关键任务', '关键职责描述模糊'),
    ]
    for p, msg in vague_patterns:
        if re.search(p, text):
            risks['vague_language']['level'] = '中'
            if msg not in risks['vague_language']['flags']:
                risks['vague_language']['flags'].append(msg)
            break

    # ── 数字完美性分析 ─
    perfect_numbers = re.findall(r'(?:100%|99%|98%|97%|96%|95%|增长了\s*\d+%|提升了\s*\d+%)', text)
    if len(perfect_numbers) > 3:
        risks['skill_exaggeration']['level'] = '中'
        risks['skill_exaggeration']['flags'].append(f"简历含{len(perfect_numbers)}处完美数据,建议核实")

    # ── 技能夸大分析 ─
    exaggeration_phrases = [
        "精通", "熟练掌握", "深度理解", "全面掌握", "专家级", "顶级",
    ]
    exp_count = 0
    for phrase in exaggeration_phrases[:4]:  # 只检查前4个
        exp_count += len(re.findall(phrase, text))

    if exp_count >= 6:
        risks['skill_exaggeration']['level'] = '高'
        risks['skill_exaggeration']['flags'].append(f'简历中{exp_count}处使用夸大技能描述')
    elif exp_count >= 4:
        risks['skill_exaggeration']['level'] = '中'
        risks['skill_exaggeration']['flags'].append(f'多处使用"精通"等夸大用词({exp_count}处)')

    # ── 工作经历风险 ─
    if len(work_exp) == 0:
        risks['work_experience']['level'] = '中'
        risks['work_experience']['flags'].append('未识别到工作经历描述')
    elif len(work_exp) < 3 and len(text) > 1500:
        risks['work_experience']['level'] = '中'
        risks['work_experience']['flags'].append('工作经历描述偏少,与简历篇幅不符')

    # ── 学历风险 ─
    if len(edu) == 0:
        risks['education']['level'] = '低'  # 不强制报警,因为PDF可能提取不全
        risks['education']['flags'].append('未识别到明确学历信息(建议人工核实)')

    # ── 计算综合风险(修复bug)──
    level_to_num = {'低': 0, '中': 1, '高': 2}
    total_score = sum(level_to_num[r['level']] for r in risks.values())
    flagged_count = sum(1 for r in risks.values() if r['flags'])

    # 综合风险:考虑flagged数量和最高风险维度
    max_risk_num = max(level_to_num[r['level']] for r in risks.values())

    overall = '低'
    if max_risk_num >= 2:  # 有高风险维度
        overall = '高'
    elif max_risk_num == 1 or flagged_count >= 3:  # 有中风险或3个以上flag
        overall = '中'

    return {
        'name': name or '未知',
        'risks': risks,
        'overall': overall,
        'summary': risk_summary(risks)
    }


def analyze_risk(resume_data):
    """Analyze resume risk with section-aware rules to reduce false positives."""
    risks = {
        "education": risk_bucket(),
        "work_experience": risk_bucket(),
        "timeline": risk_bucket(),
        "salary": risk_bucket(),
        "skill_exaggeration": risk_bucket(),
        "vague_language": risk_bucket(),
        "professional_match": risk_bucket(),
    }

    text = resume_data.get("clean_text", "")
    raw_text = resume_data.get("raw_text", "")
    pre_project_text = resume_data.get("text_before_project", raw_text)
    work_exp = resume_data.get("work_experience", [])
    edu = resume_data.get("education", [])
    skills = resume_data.get("skills", [])
    name = resume_data.get("name", "")
    parse_quality = resume_data.get("parse_quality") or assess_parse_quality(resume_data)

    non_cs_majors = ["会计", "财务", "市场", "营销", "材料", "高分子", "工艺", "工商"]
    cs_majors = ["计算机", "软件", "信息", "大数据", "统计", "数学"]
    edu_text = " ".join(edu)
    for major in non_cs_majors:
        if major in text[:500] or major in edu_text[:400]:
            add_risk(
                risks,
                "professional_match",
                "中",
                f'学历专业含"{major}",与数据开发岗位关联度偏低',
                matching_evidence(raw_text, [major]) or edu_text[:180],
            )
            break
    if any(major in text[:500] or major in edu_text[:400] for major in cs_majors):
        risks["professional_match"] = risk_bucket()

    all_work_ranges = extract_date_ranges(work_exp)
    work_ranges = [
        item for item in all_work_ranges
        if not looks_like_project_timeline_line(item[2])
    ]
    if len(work_ranges) >= 2:
        ordered_ranges = sorted(work_ranges, key=lambda item: item[0])
        latest_end = ordered_ranges[0][1]
        for start, end, line in ordered_ranges[1:]:
            overlap = latest_end - start
            if overlap >= 3:
                add_risk(risks, "timeline", "中", f"工作时间段可能重叠:{line[:60]}", line)
                break
            latest_end = max(latest_end, end)

    vague_patterns = [
        (r"某知名|某大型|头部企业|行业龙头", "公司描述使用了模糊标签"),
        (r"重要项目|核心产品|关键任务", "职责描述偏概括,建议人工核实"),
    ]
    vague_text = "\n".join(work_exp) if work_exp else pre_project_text
    for pattern, message in vague_patterns:
        match = re.search(pattern, vague_text)
        if match:
            add_risk(risks, "vague_language", "中", message, matching_evidence(vague_text, [pattern]) or match.group(0))
            break

    perfect_numbers = re.findall(r"(?:100%|99%|98%|97%|96%|95%|增长至?\s*\d+%|提升至?\s*\d+%)", text)
    if len(perfect_numbers) > 3:
        add_risk(
            risks,
            "skill_exaggeration",
            "中",
            f"简历含 {len(perfect_numbers)} 处过于完美的数字表达",
            matching_evidence(raw_text, perfect_numbers[:5]),
        )

    # ── 技能夸大分析（v4: 上下文感知，避免误判）──
    # 不再简单统计关键词次数，而是结合工作年限、技能与项目匹配度综合判断
    exaggeration_strong = ["精通", "专家级", "顶级", "全面掌握"]  # 强夸大词
    exaggeration_moderate = ["熟练掌握", "深度理解"]  # 中等词，需结合上下文

    strong_count = sum(len(re.findall(phrase, text)) for phrase in exaggeration_strong)
    moderate_count = sum(len(re.findall(phrase, text)) for phrase in exaggeration_moderate)

    # 提取工作年限
    work_years = 0
    exp_match = re.search(r"(\d{1,2})\+?\s*年(?:工作|经验)", text)
    if exp_match:
        work_years = int(exp_match.group(1))
    elif work_ranges:
        # 从工作时间段推算
        total_months = sum(end - start for start, end, _ in work_ranges)
        work_years = total_months // 12

    # 技能是否在项目经历中有体现（有项目证据支撑则更可信）
    project_text = " ".join(resume_data.get("project_experience", []))
    skill_keywords_in_projects = 0
    tech_keywords = ["sql", "python", "java", "hive", "spark", "flink", "etl",
                     "hadoop", "kettle", "sqoop", "oracle", "mysql", "bi",
                     "flume", "kafka", "doris", "clickhouse", "datax"]
    for kw in tech_keywords:
        if kw.lower() in project_text.lower():
            skill_keywords_in_projects += 1

    # 综合判定逻辑：
    # - 有5+年经验，允许较多"精通""熟练掌握"
    # - 项目经历中有技术关键词支撑，降低风险
    # - 只在极端不匹配时标记（如2年经验7+处精通）
    adjusted_strong = strong_count
    adjusted_moderate = moderate_count

    # 经验年限减免：每2年经验允许1个强夸大词
    years_allowance = work_years // 2
    adjusted_strong = max(0, strong_count - years_allowance)

    # 项目证据减免：每3个项目关键词减免1个中等词
    project_allowance = skill_keywords_in_projects // 3
    adjusted_moderate = max(0, moderate_count - project_allowance)

    # 最终判定
    total_adjusted = adjusted_strong * 2 + adjusted_moderate  # 强词权重x2
    exaggeration_evidence = matching_evidence(raw_text, exaggeration_strong + exaggeration_moderate)
    if total_adjusted >= 6 and work_years < 3:
        add_risk(
            risks,
            "skill_exaggeration",
            "高",
            f"{work_years}年经验但技能描述过度自信（强夸大词{strong_count}处，中等词{moderate_count}处）",
            exaggeration_evidence,
        )
    elif total_adjusted >= 4 and work_years < 5:
        add_risk(
            risks,
            "skill_exaggeration",
            "中",
            f"{work_years}年经验，技能措辞偏强（强夸大词{strong_count}处，中等词{moderate_count}处），项目支撑{skill_keywords_in_projects}项技术",
            exaggeration_evidence,
        )
    # 其余情况不标记技能夸大（经验足够或措辞适度）

    if len(work_exp) == 0:
        add_risk(
            risks,
            "work_experience",
            "中",
            "未识别到明确的工作经历时间线",
            "；".join(parse_quality.get("warnings", [])) or raw_text[:180],
        )
    elif len(all_work_ranges) == 0 and len(text) > 1500:
        add_risk(
            risks,
            "work_experience",
            "中",
            "工作经历存在,但缺少可识别的时间段",
            " | ".join(work_exp[:3]),
        )

    if len(edu) == 0:
        add_risk(risks, "education", "低", "未识别到明确学历信息,建议人工复核原始简历", raw_text[:180])

    skill_text = " ".join(skills)
    if skill_text and not re.search(r"(sql|python|java|hive|spark|flink|etl|bi|数据)", skill_text, re.IGNORECASE):
        if risks["professional_match"]["level"] == "低":
            risks["professional_match"]["level"] = "中"
        if not risks["professional_match"]["flags"]:
            add_risk(
                risks,
                "professional_match",
                "中",
                "技能区未识别到明显的数据开发关键词",
                " | ".join(skills[:3]),
            )

    level_to_num = {"低": 0, "中": 1, "高": 2}
    flagged_count = sum(1 for risk in risks.values() if risk["flags"])
    max_risk_num = max(level_to_num[risk["level"]] for risk in risks.values())

    overall = "低"
    if max_risk_num >= 2:
        overall = "高"
    elif max_risk_num == 1 or flagged_count >= 3:
        overall = "中"

    return {
        "name": name or "未知",
        "filename": resume_data.get("filename", ""),
        "parse_quality": parse_quality,
        "risks": risks,
        "overall": overall,
        "summary": risk_summary(risks),
    }


def create_excel_report(results, output_path):
    """生成Excel风险评估报告"""
    wb = Workbook()

    # ── Sheet 1: 风险总览 ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = '风险总览'

    headers = [
        '序号', '姓名', '文件名', '文本长度', '解析质量', '整体风险',
        '学历', '经历', '时间线', '薪资', '技能夸大', '表述模糊', '专业匹配', '风险点摘要'
    ]
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, r in enumerate(results, 1):
        parse_quality = r.get('parse_quality', {})
        row = [
            idx,
            r['name'],
            r.get('filename', ''),
            parse_quality.get('text_length', ''),
            parse_quality.get('quality', ''),
            r['overall'],
            r['risks']['education']['level'],
            r['risks']['work_experience']['level'],
            r['risks']['timeline']['level'],
            r['risks']['salary']['level'],
            r['risks']['skill_exaggeration']['level'],
            r['risks']['vague_language']['level'],
            r['risks'].get('professional_match', {}).get('level', '低'),
            r['summary']
        ]
        for col, val in enumerate(row, 1):
            cell = ws1.cell(row=idx+1, column=col, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            if col in [6, 7, 8, 9, 10, 11, 12, 13] and val in RISK_COLORS:
                cell.fill = PatternFill('solid', fgColor=RISK_COLORS[val])
                cell.font = Font(color=RISK_FONTS[val])

    widths = [6, 12, 28, 10, 10, 8, 8, 8, 8, 8, 8, 8, 8, 50]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[1].height = 25

    # ── Sheet 2: 详细分析 ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('详细分析')

    headers2 = ['序号', '姓名', '风险维度', '风险等级', '具体风险点', '触发原文']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    for idx, r in enumerate(results, 1):
        for dim, data in r['risks'].items():
            if data['flags']:
                evidence_items = data.get('evidence', [])
                for flag_index, flag in enumerate(data['flags']):
                    ws2.cell(row=row_num, column=1, value=idx)
                    ws2.cell(row=row_num, column=2, value=r['name'])
                    ws2.cell(row=row_num, column=3, value=RISK_DIMENSION_LABELS.get(dim, dim))
                    ws2.cell(row=row_num, column=4, value=data['level'])
                    ws2.cell(row=row_num, column=5, value=flag)
                    evidence = evidence_items[flag_index] if flag_index < len(evidence_items) else ''
                    ws2.cell(row=row_num, column=6, value=evidence)

                    cell = ws2.cell(row=row_num, column=4)
                    if data['level'] in RISK_COLORS:
                        cell.fill = PatternFill('solid', fgColor=RISK_COLORS[data['level']])
                        cell.font = Font(color=RISK_FONTS[data['level']])

                    row_num += 1

    for i, w in enumerate([6, 12, 15, 8, 60, 80], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: 汇总统计 ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet('汇总统计')

    ws3['A1'] = '简历风险评估汇总'
    ws3['A1'].font = Font(bold=True, size=16)

    ws3['A3'] = '风险等级分布'
    ws3['A3'].font = Font(bold=True, size=12)

    risk_counts = {'高': 0, '中': 0, '低': 0}
    for r in results:
        risk_counts[r['overall']] += 1

    ws3['A5'] = '风险等级'
    ws3['B5'] = '人数'
    ws3['C5'] = '占比'
    for cell in [ws3['A5'], ws3['B5'], ws3['C5']]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9E1F2')

    for i, (level, count) in enumerate(risk_counts.items(), 6):
        ws3.cell(row=i, column=1, value=level)
        ws3.cell(row=i, column=2, value=count)
        ws3.cell(row=i, column=3, value=f'=B{i}/SUM(B$6:B$8)')
        ws3.cell(row=i, column=3).number_format = '0.0%'

        cell = ws3.cell(row=i, column=1)
        cell.fill = PatternFill('solid', fgColor=RISK_COLORS[level])
        cell.font = Font(color=RISK_FONTS[level])

    # 高风险名单
    ws3['A11'] = '高风险候选人'
    ws3['A11'].font = Font(bold=True, size=12)
    ws3['A12'] = '序号'
    ws3['B12'] = '姓名'
    ws3['C12'] = '主要风险点'
    for cell in [ws3['A12'], ws3['B12'], ws3['C12']]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9E1F2')

    hi_risk = [r for r in results if r['overall'] == '高']
    for i, r in enumerate(hi_risk, 13):
        ws3.cell(row=i, column=1, value=i-12)
        ws3.cell(row=i, column=2, value=r['name'])
        ws3.cell(row=i, column=3, value=r['summary'][:60])

    # 中风险名单
    mid_risk = [r for r in results if r['overall'] == '中']
    if mid_risk:
        start_row = 13 + len(hi_risk) + 2
        ws3.cell(row=start_row, column=1, value='中风险候选人')
        ws3.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        ws3.cell(row=start_row+1, column=1, value='序号')
        ws3.cell(row=start_row+1, column=2, value='姓名')
        ws3.cell(row=start_row+1, column=3, value='主要风险点')
        for cell in [ws3.cell(row=start_row+1, column=1), ws3.cell(row=start_row+1, column=2), ws3.cell(row=start_row+1, column=3)]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='D9E1F2')
        for i, r in enumerate(mid_risk, start_row+2):
            ws3.cell(row=i, column=1, value=i-start_row-1)
            ws3.cell(row=i, column=2, value=r['name'])
            ws3.cell(row=i, column=3, value=r['summary'][:60])

    for i, w in enumerate([8, 15, 60], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    return output_path


def scan_resumes_folder(folder_path, recursive=True):
    """扫描文件夹中的所有简历文件,支持递归遍历子文件夹"""
    supported_extensions = ['.pdf', '.docx', '.doc', '.xlsx', '.xls', '.txt']
    ignored_name_keywords = ['风险评估报告', 'resume_risk_report', '\u63a8\u8350\u8868']
    resumes = []

    folder = Path(folder_path)
    if not folder.exists():
        return []

    if recursive:
        # 递归遍历所有子文件夹
        for file_path in folder.rglob('*'):
            if file_path.is_file():
                ext = file_path.suffix.lower()
                if ext in supported_extensions:
                    if file_path.name.startswith('~$'):
                        continue
                    if any(keyword in file_path.stem for keyword in ignored_name_keywords):
                        continue
                    resumes.append({
                        'filepath': str(file_path),
                        'filename': file_path.name
                    })
    else:
        # 只扫描当前文件夹
        for ext in supported_extensions:
            for file_path in folder.glob(f'*{ext}'):
                if file_path.name.startswith('~$'):
                    continue
                if any(keyword in file_path.stem for keyword in ignored_name_keywords):
                    continue
                resumes.append({
                    'filepath': str(file_path),
                    'filename': file_path.name
                })

    return resumes


def process_resume_file(filepath, filename=None):
    """处理单个简历文件,返回文本、姓名和文件名。"""
    text = extract_resume_text(filepath)
    resolved_filename = filename or Path(filepath).name
    name = Path(resolved_filename).stem

    # 优先从文件名末尾提取候选人姓名。
    name_match = re.search(r'[-_]([^-_]{2,4})(?:\.[^.]+)$', resolved_filename, re.IGNORECASE)
    if name_match:
        name = name_match.group(1)

    return {
        'text': text,
        'name': name,
        'filename': resolved_filename
    }


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description='简历风险评估工具 v3.0')
    parser.add_argument('--file', '-f', help='单个简历文件路径')
    parser.add_argument('--folder', '-d', help='简历文件夹路径')
    parser.add_argument('--json', '-j', help='JSON格式的简历数据')
    parser.add_argument('--output', '-o', default=None, help='输出文件路径')
    parser.add_argument('--no-recursive', action='store_true', help='仅扫描指定文件夹,不递归子文件夹')
    parser.add_argument('--debug-dir', help='保存抽取原文和解析分区诊断文件的目录')
    parser.add_argument(
        '--save-extracted-text',
        nargs='?',
        const='extraction_debug',
        help='保存每份简历的抽取文本和章节分区诊断文件；可选指定输出目录'
    )

    args = parser.parse_args()

    resumes_data = []

    # 处理不同输入方式
    if args.json:
        try:
            data = json.loads(args.json)
            if isinstance(data, list):
                resumes_data = data
            else:
                resumes_data = [data]
        except json.JSONDecodeError as e:
            print(json.dumps({'success': False, 'error': f'Invalid JSON: {e}'}))
            sys.exit(1)

    elif args.file:
        resume = process_resume_file(args.file)
        resumes_data = [resume]

    elif args.folder:
        files = scan_resumes_folder(args.folder, recursive=not args.no_recursive)
        if not files:
            print(json.dumps({'success': False, 'error': '未找到简历文件'}))
            sys.exit(1)
        for f in files:
            resume = process_resume_file(f['filepath'], f['filename'])
            resumes_data.append(resume)

    else:
        # 从标准输入读取
        input_data = sys.stdin.buffer.read()
        if input_data:
            try:
                data = json.loads(input_data.decode('utf-8'))
                if isinstance(data, list):
                    resumes_data = data
                else:
                    resumes_data = [data]
            except json.JSONDecodeError as e:
                print(json.dumps({'success': False, 'error': f'Invalid JSON: {e}'}))
                sys.exit(1)
        else:
            parser.print_help()
            sys.exit(0)

    # 解析和分析简历
    results = []
    debug_dir = args.debug_dir or args.save_extracted_text
    debug_files = []
    for index, resume in enumerate(resumes_data, 1):
        text = resume.get('text', resume.get('content', '')) if isinstance(resume, dict) else str(resume)
        name = resume.get('name') if isinstance(resume, dict) else None
        filename = resume.get('filename') if isinstance(resume, dict) else None
        parsed = parse_resume(text, name, filename)
        if debug_dir:
            debug_files.append(write_extraction_debug_files(parsed, debug_dir, index))
        analyzed = analyze_risk(parsed)
        results.append(analyzed)

    # 生成Excel报告
    if args.output:
        output_path = args.output
    elif args.folder:
        output_path = str(Path(args.folder) / 'resume_risk_report.xlsx')
    elif args.file:
        output_path = str(Path(args.file).parent / 'resume_risk_report.xlsx')
    else:
        output_path = 'resume_risk_report.xlsx'
    if HAS_OPENPYXL:
        create_excel_report(results, output_path)

    # 输出JSON结果
    output = {
        'success': True,
        'output': output_path,
        'count': len(results),
        'high_risk': len([r for r in results if r['overall'] == '高']),
        'medium_risk': len([r for r in results if r['overall'] == '中']),
        'low_risk': len([r for r in results if r['overall'] == '低']),
        'debug_dir': str(Path(debug_dir)) if debug_dir else None,
        'debug_files': debug_files,
        'details': [
            {
                'name': r['name'],
                'overall': r['overall'],
                'summary': r['summary'],
                'parse_quality': r.get('parse_quality', {}).get('quality'),
            }
            for r in results
        ]
    }

    print(json.dumps(output, ensure_ascii=False))
    return output


if __name__ == '__main__':
    main()


