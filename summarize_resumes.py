# -*- coding: utf-8 -*-
import json, os, sys, re
sys.stdout.reconfigure(encoding='utf-8')

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open(r'C:\Users\EDY\resume_risk_report\resumes_for_analysis.json', 'r', encoding='utf-8') as f:
    resumes = json.load(f)

def deep_clean(text):
    """Clean PDF extracted text"""
    # Remove watermark hex strings
    text = re.sub(r'[a-f0-9]{32,}', '', text)
    # Normalize Chinese spacing
    text = re.sub(r'([\u4e00-\u9fa5])\s+([\u4e00-\u9fa5])', r'\1\2', text)
    text = re.sub(r'([a-zA-Z0-9])\s+([a-zA-Z0-9])', r'\1\2', text)
    text = re.sub(r'\n{3,}', '\n', text)
    return text.strip()

def get_section_lines(text, keywords, max_lines=50):
    """Get lines belonging to a section"""
    lines = text.split('\n')
    result = []
    in_section = False
    section_end_keywords = ['教育', '工作', '项目', '自我', '求职', '简历', '证书', '荣誉', '语言']
    for line in lines:
        clean = line.strip()
        if not clean:
            continue
        # Check if this line starts a new section
        if any(k in clean[:6] for k in ['基本信息', '个人简历', '工 作 年 限', '工作性质', '目标职能', '个人技能']):
            if any(kw in clean for kw in ['工作经历', '项目经验', '技能特长', '工作经历']):
                in_section = True
                continue
        if in_section:
            # Stop if we hit another major section
            if any(k in clean[:6] for k in ['工作经历', '项目经验', '技能特长', '教育背景', '求职意向', '简历', '证书']):
                if len(clean) < 15:
                    break
        if in_section:
            result.append(clean)
    return result

def parse_resume(text, filename):
    text = deep_clean(text)
    name = filename.replace('深德科-数据开发-', '').replace('.pdf', '')
    
    lines = text.split('\n')
    
    # --- Basic Info ---
    info = {'姓名': name}
    
    # Gender
    info['性别'] = '男' if any('男' in l[:30] for l in lines[:15]) else ('女' if any('女' in l[:30] for l in lines[:15]) else '')
    
    # Age - look for 年龄 or just after name
    age_m = re.search(r'年龄[：:\s]*(\d{2})', text)
    if age_m:
        info['年龄'] = age_m.group(1)
    else:
        # Try finding 岁 pattern
        age_m = re.search(r'(\d{2})\s*岁', text[:300])
        info['年龄'] = age_m.group(1) if age_m else ''
    
    # Phone
    phone_m = re.search(r'1[3-9]\d[\s\-]?\d{4}[\s\-]?\d{4}', text)
    info['手机'] = phone_m.group().replace(' ', '').replace('-', '') if phone_m else ''
    
    # Email
    email_m = re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', text)
    info['邮箱'] = email_m.group() if email_m else ''
    
    # Work years
    wy_m = re.search(r'(?:工作|从业)\s*年\s*限[：:\s]*(\d+)\s*年', text)
    info['工作年限'] = (wy_m.group(1) + '年') if wy_m else ''
    
    # Target position
    tp_m = re.search(r'目标职能[：:\s]*([^\n]{5,40})', text)
    if tp_m:
        info['目标岗位'] = re.sub(r'\s+', '', tp_m.group(1))[:40]
    else:
        info['目标岗位'] = ''
    
    # --- Education ---
    edu_results = []
    for i, line in enumerate(lines):
        clean = re.sub(r'\s+', '', line)
        if any(k in clean for k in ['大学', '学院', '硕士', '博士', '本科', '大专', '一本', '二本', '985', '211']) and len(clean) > 6:
            # Skip if it's in skills section
            if i > 0:
                prev = re.sub(r'\s+', '', lines[i-1])
                if any(k in prev for k in ['技能', '技术', '专长', '擅长']):
                    continue
            edu_results.append(clean[:50])
    info['学历信息'] = ' | '.join(edu_results[:2])
    
    # --- Companies ---
    company_results = []
    for line in lines:
        clean = re.sub(r'\s+', '', line)
        m = re.search(r'(\d{4}[./\-]\d{1,2}\s*[-~–]\s*\d{4}[./\-]?\d{0,2})\s*([\u4e00-\u9fa5a-zA-Z0-9（）\(\)]{4,30}(?:公司|集团|企业|有限|科技|信息|技术|研究所))', clean)
        if m:
            company_results.append(clean[:70])
    info['工作经历'] = ' | '.join(company_results[:3])
    
    # --- Skills ---
    skill_section = False
    skill_results = []
    skip_keywords = ['工作经历', '项目经验', '教育背景', '求职意向', '自我评价', '简历', '证书', '语言能力']
    for line in lines:
        clean = re.sub(r'\s+', '', line)
        if any(k in clean[:8] for k in ['技能特长', '个人技能', '专业技能', '技术专长']) and len(clean) < 20:
            skill_section = True
            continue
        if skill_section:
            if any(k in clean[:6] for k in skip_keywords) and len(clean) < 15:
                break
            if len(clean) > 5:
                skill_results.append(clean[:45])
    info['技能特长'] = ' '.join(skill_results[:10])
    
    return info

# Parse all resumes
all_info = []
for r in resumes:
    info = parse_resume(r['text'], r['name'])
    all_info.append(info)
    print(f"  {info['姓名']} | {info['年龄']}岁 | {info['工作年限']} | {info['手机']}")

# Create DataFrame
df = pd.DataFrame(all_info)
output_path = r'C:\Users\EDY\resume_risk_report\简历信息汇总.xlsx'

# Reorder columns
col_order = ['姓名', '性别', '年龄', '工作年限', '手机', '邮箱', '目标岗位', '学历信息', '工作经历', '技能特长']
df = df[[c for c in col_order if c in df.columns]]

df.to_excel(output_path, index=False, sheet_name='简历信息汇总')

# Style workbook
wb = load_workbook(output_path)
ws = wb.active

header_fill = PatternFill('solid', fgColor='4472C4')
header_font = Font(bold=True, color='FFFFFF', size=11)
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

col_widths = [10, 6, 6, 8, 15, 25, 22, 40, 50, 60]
for i, w in enumerate(col_widths, 1):
    if i <= ws.max_column:
        ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 25
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 70

for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row=row, column=col).border = border

wb.save(output_path)
print(f"\n汇总完成: {output_path}")
